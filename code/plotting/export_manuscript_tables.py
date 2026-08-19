import argparse
from pathlib import Path

import numpy as np
import pandas as pd
from scipy.stats import norm
from sklearn.metrics import confusion_matrix, roc_auc_score
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PACKAGE_ROOT = Path(__file__).resolve().parents[2]


COHORT_COLUMNS = [
    "Training Cohort",
    "Internal Validation",
    "External Center 1",
    "External Center 2",
    "External Combined",
    "Prospective Cohort 410",
]

COHORT_LABELS = {
    "Training Cohort": "Training Cohort",
    "Internal Validation": "Internal Validation",
    "External Center 1": "External Center 1",
    "External Center 2": "External Center 2",
    "External Combined": "External Combined",
    "Prospective Cohort 410": "Prospective Cohort",
}

METRIC_ROWS = [
    ("AUC", "auc", False),
    ("Sensitivity (%)", "sensitivity", True),
    ("Specificity (%)", "specificity", True),
    ("Accuracy (%)", "accuracy", True),
    ("Balanced Acc. (%)", "balanced_acc", True),
    ("PPV (%)", "ppv", True),
    ("NPV (%)", "npv", True),
]

COMPARISON_METHODS = [
    ("Proposed AI Model (NCCT)", "AI Model"),
    ("Non-gated Agatston Score (NCCT)", "Non-gated Agatston"),
    ("Gated Agatston Score (Dedicated CSCT)", "Gated Agatston"),
]


def sensitivity(y_true, y_pred):
    tn, fp, fn, tp = confusion_matrix(y_true, y_pred, labels=[0, 1]).ravel()
    return tp / (tp + fn) if (tp + fn) else np.nan


def specificity(y_true, y_pred):
    tn, fp, fn, tp = confusion_matrix(y_true, y_pred, labels=[0, 1]).ravel()
    return tn / (tn + fp) if (tn + fp) else np.nan


def npv(y_true, y_pred):
    tn, fp, fn, tp = confusion_matrix(y_true, y_pred, labels=[0, 1]).ravel()
    return tn / (tn + fn) if (tn + fn) else np.nan


def format_estimate(value, low, high, percent=False):
    if percent:
        return f"{value * 100:.1f} ({low * 100:.1f} - {high * 100:.1f})"
    return f"{value:.3f} ({low:.3f} - {high:.3f})"


def format_pvalue(p):
    if not np.isfinite(p):
        return "p = NA"
    if p < 0.05:
        return "p < 0.05"
    return "ns"


def format_pvalue_excel(p):
    if not np.isfinite(p):
        return "p = NA"
    return f"p = {p:.6g} ({format_pvalue(p)})"


def delong_auc_variance(y_true, y_score):
    y_true = np.asarray(y_true)
    y_score = np.asarray(y_score)
    n_pos = np.sum(y_true == 1)
    n_neg = np.sum(y_true == 0)
    pos_scores = y_score[y_true == 1]
    neg_scores = y_score[y_true == 0]
    theta = 0.0
    for p in pos_scores:
        theta += np.sum(neg_scores < p) + 0.5 * np.sum(neg_scores == p)
    theta /= n_pos * n_neg

    var_pos = 0.0
    for p in pos_scores:
        v10 = (np.sum(neg_scores < p) + 0.5 * np.sum(neg_scores == p)) / n_neg
        var_pos += (v10 - theta) ** 2
    var_pos /= (n_pos - 1) if n_pos > 1 else 1

    var_neg = 0.0
    for n in neg_scores:
        v01 = (np.sum(pos_scores > n) + 0.5 * np.sum(pos_scores == n)) / n_pos
        var_neg += (v01 - theta) ** 2
    var_neg /= (n_neg - 1) if n_neg > 1 else 1
    return theta, var_pos / n_pos + var_neg / n_neg


def delong_test(y_true, score1, score2):
    y_true = np.asarray(y_true)
    score1 = np.asarray(score1)
    score2 = np.asarray(score2)
    auc1, var1 = delong_auc_variance(y_true, score1)
    auc2, var2 = delong_auc_variance(y_true, score2)
    n_pos = np.sum(y_true == 1)
    n_neg = np.sum(y_true == 0)
    pos1, neg1 = score1[y_true == 1], score1[y_true == 0]
    pos2, neg2 = score2[y_true == 1], score2[y_true == 0]

    cov_pos = 0.0
    for i in range(n_pos):
        v10_1 = (np.sum(neg1 < pos1[i]) + 0.5 * np.sum(neg1 == pos1[i])) / n_neg
        v10_2 = (np.sum(neg2 < pos2[i]) + 0.5 * np.sum(neg2 == pos2[i])) / n_neg
        cov_pos += (v10_1 - auc1) * (v10_2 - auc2)
    cov_pos = cov_pos / (n_pos - 1) if n_pos > 1 else 0

    cov_neg = 0.0
    for i in range(n_neg):
        v01_1 = (np.sum(pos1 > neg1[i]) + 0.5 * np.sum(pos1 == neg1[i])) / n_pos
        v01_2 = (np.sum(pos2 > neg2[i]) + 0.5 * np.sum(pos2 == neg2[i])) / n_pos
        cov_neg += (v01_1 - auc1) * (v01_2 - auc2)
    cov_neg = cov_neg / (n_neg - 1) if n_neg > 1 else 0

    cov_total = cov_pos / n_pos + cov_neg / n_neg
    se = np.sqrt(var1 + var2 - 2 * cov_total)
    z = (auc1 - auc2) / se if se > 0 else 0
    return 2 * (1 - norm.cdf(abs(z)))


def mcnemar_exact_pvalue(y_true, pred_a, pred_b, mask=None):
    if mask is None:
        mask = np.ones(len(y_true), dtype=bool)
    y = y_true[mask]
    a = pred_a[mask]
    b = pred_b[mask]
    correct_a = a == y
    correct_b = b == y
    only_a = int(np.sum(correct_a & ~correct_b))
    only_b = int(np.sum(~correct_a & correct_b))
    discordant = only_a + only_b
    if discordant == 0:
        return 1.0
    # Exact two-sided binomial test without importing scipy.stats.binomtest for older envs.
    from scipy.stats import binomtest

    return float(binomtest(min(only_a, only_b), discordant, 0.5, alternative="two-sided").pvalue)


def generate_bootstrap_indices(n, centers=None, seed_rng=None):
    if seed_rng is None:
        seed_rng = np.random.default_rng()
    idx_all = np.arange(n)
    if centers is None:
        return seed_rng.choice(idx_all, size=n, replace=True)

    centers = np.asarray(centers)
    sampled = []
    for center in np.unique(centers):
        center_idx = idx_all[centers == center]
        sampled.append(seed_rng.choice(center_idx, size=len(center_idx), replace=True))
    return np.concatenate(sampled)


def bootstrap_metric_diff_ci_p(y_true, pred1, pred2, metric_func, centers=None, n_bootstrap=10000, seed=42):
    rng = np.random.default_rng(seed)
    point = metric_func(y_true, pred1) - metric_func(y_true, pred2)
    if not np.isfinite(point):
        return point, np.nan, np.nan, np.nan

    diffs = []
    for _ in range(n_bootstrap):
        idx = generate_bootstrap_indices(len(y_true), centers=centers, seed_rng=rng)
        v1 = metric_func(y_true[idx], pred1[idx])
        v2 = metric_func(y_true[idx], pred2[idx])
        if np.isfinite(v1) and np.isfinite(v2):
            diffs.append(v1 - v2)
    if not diffs:
        return point, np.nan, np.nan, np.nan

    diffs = np.asarray(diffs)
    low, high = np.percentile(diffs, [2.5, 97.5]).astype(float)
    null_centered = diffs - point
    p_value = (np.sum(np.abs(null_centered) >= abs(point)) + 1) / (len(diffs) + 1)
    return point, low, high, min(float(p_value), 1.0)


def build_table1(metrics_df):
    rows = []
    for label, key, percent in METRIC_ROWS:
        row = {"Metrics": label}
        for cohort in COHORT_COLUMNS:
            data = metrics_df.loc[metrics_df["cohort"] == cohort].iloc[0]
            header = f"{COHORT_LABELS[cohort]} (n={int(data['n'])})"
            row[header] = format_estimate(data[key], data[f"ci_{key}_low"], data[f"ci_{key}_high"], percent=percent)
        rows.append(row)
    return pd.DataFrame(rows)


def align_ai_metrics_with_comparison(metrics_df, calcium_metrics):
    """Use the Table 2 aligned comparison AI rows for overlapping cohorts."""
    aligned = metrics_df.copy()
    column_map = {
        "n": "N",
        "tn": "TN",
        "fp": "FP",
        "fn": "FN",
        "tp": "TP",
        "auc": "AUC",
        "sensitivity": "Sensitivity",
        "specificity": "Specificity",
        "accuracy": "Accuracy",
        "balanced_acc": "Balanced Accuracy",
        "ppv": "PPV",
        "npv": "NPV",
        "ci_auc_low": "AUC 95% CI Low",
        "ci_auc_high": "AUC 95% CI High",
        "ci_sensitivity_low": "Sensitivity 95% CI Low",
        "ci_sensitivity_high": "Sensitivity 95% CI High",
        "ci_specificity_low": "Specificity 95% CI Low",
        "ci_specificity_high": "Specificity 95% CI High",
        "ci_accuracy_low": "Accuracy 95% CI Low",
        "ci_accuracy_high": "Accuracy 95% CI High",
        "ci_balanced_acc_low": "Balanced Accuracy 95% CI Low",
        "ci_balanced_acc_high": "Balanced Accuracy 95% CI High",
        "ci_ppv_low": "PPV 95% CI Low",
        "ci_ppv_high": "PPV 95% CI High",
        "ci_npv_low": "NPV 95% CI Low",
        "ci_npv_high": "NPV 95% CI High",
    }
    for cohort in ["External Combined", "Prospective Cohort 410"]:
        source = calcium_metrics[
            (calcium_metrics["Cohort"].astype(str) == cohort)
            & (calcium_metrics["Method"].astype(str) == "AI Model")
        ]
        target_idx = aligned.index[aligned["cohort"].astype(str) == cohort]
        if source.empty or len(target_idx) == 0:
            continue
        source_row = source.iloc[0]
        idx = target_idx[0]
        for out_col, source_col in column_map.items():
            if out_col in aligned.columns and source_col in source_row:
                aligned.loc[idx, out_col] = source_row[source_col]
    return aligned


def method_values(metric_df, cohort, method, metric_label):
    row = metric_df[(metric_df["Cohort"] == cohort) & (metric_df["Method"] == method)].iloc[0]
    low_key = f"{metric_label} 95% CI Low"
    high_key = f"{metric_label} 95% CI High"
    percent = metric_label != "AUC"
    return format_estimate(row[metric_label], row[low_key], row[high_key], percent=percent)


def cohort_arrays(case_df):
    y_true = case_df["GT Label"].astype(int).to_numpy()
    non_score = case_df["Non-gated"].astype(float).to_numpy()
    gated_score = case_df["Gated"].astype(float).to_numpy()
    ai_score = case_df["Pred Score"].astype(float).to_numpy()
    return {
        "y_true": y_true,
        "Non-gated Agatston": {"score": non_score, "pred": (non_score >= 100).astype(int)},
        "Gated Agatston": {"score": gated_score, "pred": (gated_score >= 100).astype(int)},
        "AI Model": {"score": ai_score, "pred": case_df["Pred Label"].astype(int).to_numpy()},
    }


def cohort_centers(case_df):
    if "Site" in case_df.columns:
        return case_df["Site"].astype(str).str.strip().to_numpy()
    return None


def comparison_values(case_df):
    arrays = cohort_arrays(case_df)
    y_true = arrays["y_true"]
    ai = arrays["AI Model"]
    non = arrays["Non-gated Agatston"]
    gated = arrays["Gated Agatston"]
    centers = cohort_centers(case_df)
    values = {
        ("AUC", "non"): format_pvalue_excel(delong_test(y_true, ai["score"], non["score"])),
        ("AUC", "gated"): format_pvalue_excel(delong_test(y_true, ai["score"], gated["score"])),
        ("Sensitivity (%)", "non"): format_pvalue_excel(mcnemar_exact_pvalue(y_true, ai["pred"], non["pred"], y_true == 1)),
        ("Sensitivity (%)", "gated"): format_pvalue_excel(mcnemar_exact_pvalue(y_true, ai["pred"], gated["pred"], y_true == 1)),
        ("Specificity (%)", "non"): format_pvalue_excel(mcnemar_exact_pvalue(y_true, ai["pred"], non["pred"], y_true == 0)),
        ("Specificity (%)", "gated"): format_pvalue_excel(mcnemar_exact_pvalue(y_true, ai["pred"], gated["pred"], y_true == 0)),
    }
    for label, comparator, pred in [
        ("NPV (%)", "non", non["pred"]),
        ("NPV (%)", "gated", gated["pred"]),
    ]:
        diff, low, high, p_value = bootstrap_metric_diff_ci_p(y_true, ai["pred"], pred, npv, centers=centers)
        values[(label, comparator)] = (
            f"Delta {diff * 100:+.1f} pp ({low * 100:+.1f} to {high * 100:+.1f}); "
            f"{format_pvalue_excel(p_value)}"
        )
    return values


def comparison_pvalue_rows(case_df, cohort_label):
    arrays = cohort_arrays(case_df)
    y_true = arrays["y_true"]
    ai = arrays["AI Model"]
    non = arrays["Non-gated Agatston"]
    gated = arrays["Gated Agatston"]
    centers = cohort_centers(case_df)
    npv_non = bootstrap_metric_diff_ci_p(y_true, ai["pred"], non["pred"], npv, centers=centers)
    npv_gated = bootstrap_metric_diff_ci_p(y_true, ai["pred"], gated["pred"], npv, centers=centers)
    specs = [
        ("AUC", "AI Model vs Non-gated Agatston", delong_test(y_true, ai["score"], non["score"]), np.nan, np.nan, np.nan),
        ("AUC", "AI Model vs Gated Agatston", delong_test(y_true, ai["score"], gated["score"]), np.nan, np.nan, np.nan),
        (
            "Sensitivity",
            "AI Model vs Non-gated Agatston",
            mcnemar_exact_pvalue(y_true, ai["pred"], non["pred"], y_true == 1),
            np.nan,
            np.nan,
            np.nan,
        ),
        (
            "Sensitivity",
            "AI Model vs Gated Agatston",
            mcnemar_exact_pvalue(y_true, ai["pred"], gated["pred"], y_true == 1),
            np.nan,
            np.nan,
            np.nan,
        ),
        (
            "Specificity",
            "AI Model vs Non-gated Agatston",
            mcnemar_exact_pvalue(y_true, ai["pred"], non["pred"], y_true == 0),
            np.nan,
            np.nan,
            np.nan,
        ),
        (
            "Specificity",
            "AI Model vs Gated Agatston",
            mcnemar_exact_pvalue(y_true, ai["pred"], gated["pred"], y_true == 0),
            np.nan,
            np.nan,
            np.nan,
        ),
        (
            "NPV",
            "AI Model vs Non-gated Agatston",
            npv_non[3],
            npv_non[0],
            npv_non[1],
            npv_non[2],
        ),
        (
            "NPV",
            "AI Model vs Gated Agatston",
            npv_gated[3],
            npv_gated[0],
            npv_gated[1],
            npv_gated[2],
        ),
    ]
    return [
        {
            "Cohort": cohort_label,
            "Metric": metric,
            "Comparison": comparison,
            "Raw p value": p_value,
            "Excel text": format_pvalue_excel(p_value),
            "Figure display": format_pvalue(p_value),
            "Diff AI-Comparator": diff,
            "Diff 95% CI Low": low,
            "Diff 95% CI High": high,
        }
        for metric, comparison, p_value, diff, low, high in specs
    ]


def build_table2(metric_df, external_cases, prospective_cases):
    rows = []
    cohorts = [
        ("External Validation Cohorts", "External Combined", external_cases),
        ("Prospective Cohort", "Prospective Cohort 410", prospective_cases),
    ]
    for section, cohort, cases in cohorts:
        rows.append({"Metrics": section})
        comparisons = comparison_values(cases)
        for metric_label in ["AUC", "Sensitivity (%)", "Specificity (%)", "NPV (%)"]:
            row = {"Metrics": metric_label}
            for out_label, method in COMPARISON_METHODS:
                row[out_label] = method_values(metric_df, cohort, method, metric_label.replace(" (%)", ""))
            row["Comparison with Non-gated Score"] = comparisons[(metric_label, "non")]
            row["Comparison with Gated Score"] = comparisons[(metric_label, "gated")]
            rows.append(row)
    return pd.DataFrame(rows)


def build_pvalue_table(external_cases, prospective_cases):
    return pd.DataFrame(
        comparison_pvalue_rows(external_cases, "External Validation Cohorts")
        + comparison_pvalue_rows(prospective_cases, "Prospective Cohort")
    )


def style_workbook(path):
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="F2F2F2")
    section_fill = PatternFill("solid", fgColor="E8EEF7")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in wb.worksheets:
        ws.freeze_panes = "B2"
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.font = Font(name="Times New Roman", size=12)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(name="Times New Roman", size=13, bold=True)
        for row_idx in range(2, ws.max_row + 1):
            value = ws.cell(row_idx, 1).value
            if value in {"External Validation Cohorts", "Prospective Cohort"}:
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row_idx, col_idx)
                    cell.fill = section_fill
                    cell.font = Font(name="Times New Roman", size=12, bold=True)
        widths = {
            "A": 24,
            "B": 34,
            "C": 36,
            "D": 40,
            "E": 34,
            "F": 34,
            "G": 34,
        }
        for idx in range(1, ws.max_column + 1):
            letter = get_column_letter(idx)
            ws.column_dimensions[letter].width = widths.get(letter, 28)
        for idx in range(1, ws.max_row + 1):
            ws.row_dimensions[idx].height = 28
    wb.save(path)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--report-dir", type=Path, default=PACKAGE_ROOT / "reports" / "case_results")
    args = parser.parse_args()
    report_dir = args.report_dir
    metrics_df_raw = pd.read_csv(report_dir / "metrics_summary_raw.csv")
    calcium_xlsx = report_dir / "paper_plots" / "calcium_ai_comparison_plot_data_aligned.xlsx"
    calcium_metrics = pd.read_excel(calcium_xlsx, sheet_name="Metrics used for plots")
    metrics_df = align_ai_metrics_with_comparison(metrics_df_raw, calcium_metrics)
    external_cases = pd.read_excel(calcium_xlsx, sheet_name="External matched cases")
    prospective_cases = pd.read_excel(calcium_xlsx, sheet_name="Prospective matched cases")

    table1 = build_table1(metrics_df)
    table2 = build_table2(calcium_metrics, external_cases, prospective_cases)
    pvalue_table = build_pvalue_table(external_cases, prospective_cases)

    output = report_dir / "manuscript_tables_ai_calcium.xlsx"
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        table1.to_excel(writer, sheet_name="Table 1 AI performance", index=False)
        table2.to_excel(writer, sheet_name="Table 2 Calcium comparison", index=False)
        pvalue_table.to_excel(writer, sheet_name="Calcium p values", index=False)
        metrics_df.to_excel(writer, sheet_name="Source AI metrics", index=False)
        calcium_metrics.to_excel(writer, sheet_name="Source calcium metrics", index=False)
        metrics_df_raw.to_excel(writer, sheet_name="Raw inference AI metrics", index=False)
    style_workbook(output)
    print(output)


if __name__ == "__main__":
    main()
