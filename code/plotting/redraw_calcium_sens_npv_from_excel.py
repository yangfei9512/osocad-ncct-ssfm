#!/usr/bin/env python3
"""Redraw calcium-score Sensitivity/NPV bar figures from the final Excel table.

The bar heights and p-value labels are read directly from the
"Table 2 Calcium comparison" sheet so the figure text matches the workbook.
"""

from __future__ import annotations

import os
import re
from pathlib import Path

os.environ.setdefault("MPLCONFIGDIR", "/tmp/matplotlib-sh-rj")

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
from PIL import Image as PILImage


PACKAGE_DIR = Path(__file__).resolve().parents[2]
CASE_RESULTS_DIR = PACKAGE_DIR / "reports" / "case_results"
FINAL_EXCEL = PACKAGE_DIR / "20260624T115247_final_all_tables_figures.xlsx"

COMPARISON_NAMES = ["Non-gated Agatston", "Gated Agatston", "AI Model"]
COMPARISON_COLORS = ["#B8C6CE", "#5C7288", "#35B9A5"]

FIGURE_ORDER = [
    "FigA_All_Cohort_ROC.png",
    "Fig1_Internal_ROC.png",
    "Fig2_External_Combined_ROC.png",
    "Fig3_Metrics_Bar.png",
    "Fig4_1_External_Center_Bars.png",
    "Fig4_2_External_Center_ROCs.png",
    "Fig4_3_Prospective_ROC.png",
    "Fig5_Scatter_Scores_Internal.png",
    "Fig5_Scatter_Scores_External.png",
    "Fig5_Scatter_Scores_Prospective.png",
    "FigC_Confusion_Matrix_Grid.png",
    "fig2c.png",
    "fig3c.png",
    "fig4c.png",
    "fig5b.png",
    "fig5c.png",
    "fig5d.png",
    "Fig6_External_Prospective_Calcium_AI_ROCs.png",
    "fig_external_calcium_ai_comparison_roc.png",
    "fig_external_calcium_ai_comparison_sens_npv.png",
    "fig_external_calcium_ai_comparison_metrics.png",
    "fig5_external_combined_roc.png",
    "fig5_external_combined_bar.png",
    "fig5_external_combined_comparison_roc.png",
    "fig5_external_combined_comparison_bar.png",
]


def parse_metric_percent(value) -> float:
    """Return the first displayed percentage value from a Table 2 cell."""
    if value is None:
        raise ValueError("Missing metric cell value.")
    match = re.search(r"[-+]?\d+(?:\.\d+)?", str(value))
    if not match:
        raise ValueError(f"Cannot parse metric value from {value!r}.")
    return float(match.group(0))


def parse_p_display(value) -> str:
    """Return the exact p-value display tier used by figures."""
    text = "" if value is None else str(value)
    if re.search(r"\bns\b", text):
        return "ns"
    less_than = re.search(r"p\s*<\s*([0-9.eE+-]+)", text)
    if less_than:
        return "p < 0.05" if float(less_than.group(1)) < 0.05 else "ns"
    match = re.search(r"p\s*=\s*([0-9.eE+-]+)", text)
    if match:
        p = float(match.group(1))
        if p < 0.05:
            return "p < 0.05"
        return "ns"
    raise ValueError(f"Cannot parse p-value display from {value!r}.")


def find_section_rows(ws, section_label: str) -> dict[str, int]:
    """Map metric names to row numbers within a Table 2 section."""
    start_row = None
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 1).value == section_label:
            start_row = row
            break
    if start_row is None:
        raise ValueError(f"Section not found: {section_label}")

    metric_rows = {}
    for row in range(start_row + 1, ws.max_row + 1):
        label = ws.cell(row, 1).value
        if label in ("External Validation Cohorts", "Prospective Cohort") and row != start_row:
            break
        if label:
            metric_rows[str(label)] = row
    return metric_rows


def load_plot_data(excel_path: Path):
    wb = load_workbook(excel_path, data_only=True)
    ws = wb["Table 2 Calcium comparison"]

    sections = [
        ("External Validation Cohorts", "External Validation Sensitivity and NPV"),
        ("Prospective Cohort", "Prospective Comparison"),
    ]
    method_cols = {
        "AI Model": 2,
        "Non-gated Agatston": 3,
        "Gated Agatston": 4,
    }

    data = {}
    for section, title in sections:
        metric_rows = find_section_rows(ws, section)
        sens_row = metric_rows["Sensitivity (%)"]
        npv_row = metric_rows["NPV (%)"]
        values = {}
        cells = {}
        for method in COMPARISON_NAMES:
            col = method_cols[method]
            values[method] = [
                parse_metric_percent(ws.cell(sens_row, col).value),
                parse_metric_percent(ws.cell(npv_row, col).value),
            ]
            cells[method] = [ws.cell(sens_row, col).coordinate, ws.cell(npv_row, col).coordinate]
        p_values = {
            "Sensitivity": {
                "AI vs Non-gated": parse_p_display(ws.cell(sens_row, 5).value),
                "AI vs Gated": parse_p_display(ws.cell(sens_row, 6).value),
                "cells": [ws.cell(sens_row, 5).coordinate, ws.cell(sens_row, 6).coordinate],
            },
            "NPV": {
                "AI vs Non-gated": parse_p_display(ws.cell(npv_row, 5).value),
                "AI vs Gated": parse_p_display(ws.cell(npv_row, 6).value),
                "cells": [ws.cell(npv_row, 5).coordinate, ws.cell(npv_row, 6).coordinate],
            },
        }
        data[section] = {"title": title, "values": values, "cells": cells, "p_values": p_values}
    return data


def beautify_axis(ax):
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_linewidth(1.0)
    ax.spines["bottom"].set_linewidth(1.0)
    ax.tick_params(axis="both", labelsize=10, width=0.8)


def draw_sig_bracket(ax, x1, x2, y, text, h=2.2):
    is_significant = text not in ("ns", "p = NA")
    text_color = "#2fb596" if is_significant else "#222222"
    text_weight = "bold" if is_significant else "normal"
    ax.plot([x1, x1, x2, x2], [y, y + h, y + h, y], lw=1.3, color="#555555", clip_on=False)
    ax.text(
        (x1 + x2) / 2,
        y + h + 1.0,
        text,
        ha="center",
        va="bottom",
        fontsize=8,
        fontweight=text_weight,
        color=text_color,
    )


def draw_bar_figure(plot_data: dict, output_path: Path):
    metric_labels = ["Sensitivity", "NPV"]
    x = [0, 1]
    width = 0.24
    fig, ax = plt.subplots(figsize=(7.7, 7.7))

    metric_values_by_group = [[], []]
    for i, (method, color) in enumerate(zip(COMPARISON_NAMES, COMPARISON_COLORS)):
        vals = plot_data["values"][method]
        for group_idx, val in enumerate(vals):
            metric_values_by_group[group_idx].append(val)
        bar_x = [group_x + (i - 1) * width for group_x in x]
        bars = ax.bar(bar_x, vals, width, label=method, color=color, edgecolor="white")
        for bar, val in zip(bars, vals):
            ax.text(
                bar.get_x() + bar.get_width() / 2,
                min(val + 1.2, 99),
                f"{val:.1f}",
                ha="center",
                va="bottom",
                fontsize=8,
                fontweight="bold",
            )

    bracket_tops = []
    for group_idx, metric_label in enumerate(metric_labels):
        group_x = x[group_idx]
        group_max = max(metric_values_by_group[group_idx])
        inner_y = group_max + 8.0
        outer_y = inner_y + 12.0
        bracket_tops.append(outer_y + 3.5)

        draw_sig_bracket(
            ax,
            group_x,
            group_x + width,
            inner_y,
            plot_data["p_values"][metric_label]["AI vs Gated"],
        )
        draw_sig_bracket(
            ax,
            group_x - width,
            group_x + width,
            outer_y,
            plot_data["p_values"][metric_label]["AI vs Non-gated"],
        )

    ax.set_xticks(x)
    ax.set_xticklabels(metric_labels)
    ax.set_ylabel("Percentage (%)")
    ax.set_ylim(0, max(112, max(bracket_tops) + 5))
    ax.set_title(plot_data["title"], fontweight="bold", y=1.16)
    ax.legend(loc="lower right", bbox_to_anchor=(1.0, 1.02), borderaxespad=0, frameon=False, fontsize=8)
    beautify_axis(ax)
    plt.tight_layout(rect=[0, 0, 1, 0.88])
    output_path.parent.mkdir(parents=True, exist_ok=True)
    plt.savefig(output_path, dpi=300, bbox_inches="tight")
    plt.close(fig)


def refresh_figures_sheet(excel_path: Path, plot_dir: Path):
    wb = load_workbook(excel_path)
    if "Figures" in wb.sheetnames:
        idx = wb.sheetnames.index("Figures")
        del wb["Figures"]
    else:
        idx = len(wb.sheetnames)
    ws = wb.create_sheet("Figures", idx)
    ws["A1"] = "Figure previews regenerated from current paper_plots PNG files"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = str(plot_dir.relative_to(PACKAGE_DIR))
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 95

    row = 4
    for name in FIGURE_ORDER:
        path = plot_dir / name
        if not path.exists():
            continue
        ws.cell(row=row, column=1, value=name).font = Font(bold=True)
        ws.cell(
            row=row + 1,
            column=1,
            value=str(path.relative_to(PACKAGE_DIR)),
        ).alignment = Alignment(wrap_text=True, vertical="top")
        img = XLImage(str(path))
        with PILImage.open(path) as im:
            width, height = im.size
        scale = min(620 / width, 360 / height, 1.0)
        img.width = int(width * scale)
        img.height = int(height * scale)
        ws.add_image(img, f"B{row}")
        ws.row_dimensions[row].height = 24
        ws.row_dimensions[row + 1].height = max(60, img.height * 0.75)
        row += 18
    wb.save(excel_path)


def rebuild_html():
    from build_github_english_release import main as build_release

    build_release()


def main():
    data = load_plot_data(FINAL_EXCEL)
    output_specs = [
        (
            "External Validation Cohorts",
            [
                "fig_external_calcium_ai_comparison_sens_npv.png",
                "fig5_external_combined_comparison_bar.png",
            ],
        ),
        ("Prospective Cohort", ["fig5b.png"]),
    ]

    for output_dir in [PACKAGE_DIR / "paper_plots", CASE_RESULTS_DIR / "paper_plots"]:
        for section, filenames in output_specs:
            for filename in filenames:
                draw_bar_figure(data[section], output_dir / filename)
                print(f"Saved {output_dir / filename}")

    refresh_figures_sheet(FINAL_EXCEL, PACKAGE_DIR / "paper_plots")
    rebuild_html()

    print("Source workbook:", FINAL_EXCEL)
    for section, section_data in data.items():
        print(section)
        for method in COMPARISON_NAMES:
            sens, npv = section_data["values"][method]
            sens_cell, npv_cell = section_data["cells"][method]
            print(f"  {method}: Sensitivity {sens:.1f} ({sens_cell}), NPV {npv:.1f} ({npv_cell})")
        for metric in ["Sensitivity", "NPV"]:
            p_info = section_data["p_values"][metric]
            print(
                f"  {metric} p: AI vs Non-gated {p_info['AI vs Non-gated']} ({p_info['cells'][0]}), "
                f"AI vs Gated {p_info['AI vs Gated']} ({p_info['cells'][1]})"
            )


if __name__ == "__main__":
    main()
