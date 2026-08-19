#!/usr/bin/env python3
"""Build English-only GitHub release data from the final Excel workbook.

The generated CSV/JSON/HTML files use the final workbook as the single source
of truth for displayed metrics.
"""

from __future__ import annotations

import csv
import html
import json
import re
import shutil
from pathlib import Path

from openpyxl import load_workbook


PACKAGE_DIR = Path(__file__).resolve().parents[2]
FINAL_EXCEL = PACKAGE_DIR / "20260624T115247_final_all_tables_figures.xlsx"
PLOT_DIR = PACKAGE_DIR / "paper_plots"
SELECTED_FIGURE_DIR = PACKAGE_DIR / "github_selected_figures"
GITHUB_DIR = PACKAGE_DIR / "github_data"
GITHUB_FIGURE_DIR = GITHUB_DIR / "figures"

TABLE_SHEETS = {
    "table1_ai_performance": "Table 1 AI performance",
    "table2_calcium_comparison": "Table 2 Calcium comparison",
}

FIGURE_ORDER = [
    "Fig2_A.png",
    "Fig2_B.png",
    "Fig2_C.png",
    "Fig3_A.png",
    "Fig3_B.png",
    "Fig3_C.png",
    "Fig3_D.png",
    "Fig4_A.png",
    "Fig4_B.png",
    "Fig4_C.png",
    "Fig4_D.png",
    "Fig5_A.png",
    "Fig5_B.png",
    "Fig5_C.png",
    "Fig5_D.png",
]

FIGURE_TITLES = {
    "Fig2_A.png": "Figure 2A",
    "Fig2_B.png": "Figure 2B",
    "Fig2_C.png": "Figure 2C",
    "Fig3_A.png": "Figure 3A",
    "Fig3_B.png": "Figure 3B",
    "Fig3_C.png": "Figure 3C",
    "Fig3_D.png": "Figure 3D",
    "Fig4_A.png": "Figure 4A",
    "Fig4_B.png": "Figure 4B",
    "Fig4_C.png": "Figure 4C",
    "Fig4_D.png": "Figure 4D",
    "Fig5_A.png": "Figure 5A",
    "Fig5_B.png": "Figure 5B",
    "Fig5_C.png": "Figure 5C",
    "Fig5_D.png": "Figure 5D",
}


def figure_source_dir() -> Path:
    if SELECTED_FIGURE_DIR.exists():
        return SELECTED_FIGURE_DIR
    return PLOT_DIR


def root_figure_prefix() -> str:
    if SELECTED_FIGURE_DIR.exists():
        return "github_selected_figures/"
    return "paper_plots/"


def cell_to_text(value) -> str:
    if value is None:
        return ""
    return str(value)


def read_display_table(workbook, sheet_name: str) -> list[list[str]]:
    ws = workbook[sheet_name]
    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        values = [cell_to_text(v) for v in row]
        while values and values[-1] == "":
            values.pop()
        if values:
            rows.append(values)
    return rows


def write_csv(path: Path, rows: list[list[str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerows(rows)


def table_to_html(rows: list[list[str]]) -> str:
    if not rows:
        return ""
    head, body = rows[0], rows[1:]
    parts = ["<table>", "<thead><tr>"]
    for value in head:
        parts.append(f"<th>{html.escape(value)}</th>")
    parts.append("</tr></thead><tbody>")
    for row in body:
        parts.append("<tr>")
        for idx, value in enumerate(row):
            tag = "th" if idx == 0 else "td"
            parts.append(f"<{tag}>{html.escape(value)}</{tag}>")
        parts.append("</tr>")
    parts.append("</tbody></table>")
    return "".join(parts)


def build_html(table_data: dict[str, list[list[str]]], figure_prefix: str, data_prefix: str) -> str:
    source_dir = figure_source_dir()
    figure_cards = []
    for name in FIGURE_ORDER:
        src = source_dir / name
        if not src.exists():
            continue
        title = FIGURE_TITLES.get(name, name)
        figure_cards.append(
            "<article class='figure-card'>"
            f"<img src='{html.escape(figure_prefix + name)}' alt='{html.escape(title)}'>"
            f"<h2>{html.escape(title)}</h2>"
            "</article>"
        )

    tables = []
    for key, rows in table_data.items():
        title = "AI Diagnostic Performance" if key == "table1_ai_performance" else "AI and Calcium Score Comparison"
        tables.append(
            "<section class='table-section'>"
            f"<h2>{html.escape(title)}</h2>"
            f"<div class='table-wrap'>{table_to_html(rows)}</div>"
            "</section>"
        )

    css = """
body{margin:0;padding:28px;background:#f7f8fa;color:#222;font-family:Arial,Helvetica,sans-serif}
h1{margin:0 0 8px;font-size:28px}
p{margin:0 0 20px;color:#555;line-height:1.45}
.grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(360px,1fr));gap:22px}
.figure-card,.table-section{background:#fff;border:1px solid #e4e6eb;border-radius:8px;padding:14px;box-shadow:0 2px 8px rgba(0,0,0,.05)}
img{width:100%;height:auto;display:block}
.figure-card h2{font-size:16px;margin:12px 0 0;text-align:center}
.table-section{margin-top:26px}
.table-section h2{margin:0 0 12px;font-size:20px}
.table-wrap{overflow:auto;border:1px solid #e5e7eb;border-radius:6px}
table{border-collapse:collapse;width:100%;font-size:13px;background:white}
th,td{border-bottom:1px solid #eceff3;border-right:1px solid #f1f3f5;padding:7px 9px;white-space:nowrap;text-align:center}
thead th{background:#f0f3f7;color:#222}
tbody th{text-align:left;background:#fbfcfd}
.downloads a{color:#1469c8;text-decoration:none;font-weight:600}
"""
    return (
        "<!doctype html><html><head><meta charset='utf-8'>"
        "<title>SH RJ AI Performance Results</title>"
        f"<style>{css}</style></head><body>"
        "<h1>SH RJ AI Performance Results</h1>"
        "<p>English GitHub release. Displayed metric values are read directly from the final Excel workbook.</p>"
        "<p class='downloads'>"
        f"<a href='{html.escape(data_prefix)}table1_ai_performance.csv'>Table 1 CSV</a> | "
        f"<a href='{html.escape(data_prefix)}table2_calcium_comparison.csv'>Table 2 CSV</a> | "
        f"<a href='{html.escape(data_prefix)}metrics_summary.json'>Metrics JSON</a>"
        "</p>"
        f"<section class='grid'>{''.join(figure_cards)}</section>"
        f"{''.join(tables)}"
        "</body></html>"
    )


def has_cjk(path: Path) -> bool:
    if path.suffix.lower() not in {".csv", ".json", ".html", ".md"}:
        return False
    text = path.read_text(encoding="utf-8", errors="ignore")
    return re.search(r"[\u4e00-\u9fff]", text) is not None


def main() -> None:
    wb = load_workbook(FINAL_EXCEL, data_only=True)
    table_data = {
        output_name: read_display_table(wb, sheet_name)
        for output_name, sheet_name in TABLE_SHEETS.items()
    }

    GITHUB_DIR.mkdir(parents=True, exist_ok=True)
    GITHUB_FIGURE_DIR.mkdir(parents=True, exist_ok=True)

    for output_name, rows in table_data.items():
        write_csv(GITHUB_DIR / f"{output_name}.csv", rows)

    metrics_json = {
        "source_excel": FINAL_EXCEL.name,
        "tables": {
            output_name: {
                "source_sheet": TABLE_SHEETS[output_name],
                "rows": rows,
            }
            for output_name, rows in table_data.items()
        },
    }
    (GITHUB_DIR / "metrics_summary.json").write_text(
        json.dumps(metrics_json, indent=2, ensure_ascii=True),
        encoding="utf-8",
    )

    for name in FIGURE_ORDER:
        src = figure_source_dir() / name
        if src.exists():
            shutil.copy2(src, GITHUB_FIGURE_DIR / name)

    github_html = build_html(table_data, "figures/", "")
    (GITHUB_DIR / "index.html").write_text(github_html, encoding="utf-8")

    root_html = build_html(table_data, root_figure_prefix(), "github_data/")
    (PACKAGE_DIR / "index.html").write_text(root_html, encoding="utf-8")
    (PACKAGE_DIR / "html1.html").write_text(root_html, encoding="utf-8")

    readme = """# SH RJ AI Performance Results

This folder contains the English GitHub release data for the SH RJ project.

- `table1_ai_performance.csv`: AI diagnostic performance, copied from the final Excel workbook.
- `table2_calcium_comparison.csv`: AI versus calcium score comparison, copied from the final Excel workbook.
- `metrics_summary.json`: JSON copy of the same displayed metric tables.
- `figures/`: selected publication figures used by the GitHub HTML page.
- `index.html`: English-only HTML summary page.

Displayed metric values are aligned to:
`20260624T115247_final_all_tables_figures.xlsx`.
"""
    (GITHUB_DIR / "README.md").write_text(readme, encoding="utf-8")

    generated = [
        GITHUB_DIR / "table1_ai_performance.csv",
        GITHUB_DIR / "table2_calcium_comparison.csv",
        GITHUB_DIR / "metrics_summary.json",
        GITHUB_DIR / "index.html",
        GITHUB_DIR / "README.md",
        PACKAGE_DIR / "index.html",
        PACKAGE_DIR / "html1.html",
    ]
    check = {
        "source_excel": FINAL_EXCEL.name,
        "generated_files": [str(p.relative_to(PACKAGE_DIR)) for p in generated],
        "alignment": "PASS: GitHub CSV/JSON/HTML tables were generated directly from the final Excel workbook.",
        "english_only_text_files": "PASS" if not any(has_cjk(p) for p in generated) else "FAIL",
    }
    (GITHUB_DIR / "alignment_check.json").write_text(
        json.dumps(check, indent=2, ensure_ascii=True),
        encoding="utf-8",
    )
    print(json.dumps(check, indent=2, ensure_ascii=True))


if __name__ == "__main__":
    main()
